#!/usr/bin/env python3
# pip install openpyxl

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font

# open file
wb = openpyxl.load_workbook('example.xlsx')
type(wb)
#* <class 'openpyxl.workbook.workbook.Workbook'>

# get sheet's names
wb.get_sheet_names()
#* ['Sheet1', 'Sheet2', 'Sheet3']

# select one sheet to work with
sheet = wb['Sheet1']
sheet
#* <Worksheet "Sheet1">
type(sheet)
#* <class 'openpyxl.worksheet.worksheet.Worksheet'>
sheet.title
#* 'Sheet1'

# get the workbook’s active sheet
anotherSheet = wb.active
anotherSheet
#* <Worksheet "Sheet1">

# working with cells
sheet['A1']
#* <Cell 'Sheet1'.A1>
sheet['A1'].value
#* '4/5/2015 1:34:02 PM'

c = sheet['B1']
c.value
#* 'Apples'

print(f'Row {c.row}, Column {c.column} is {c.value}')
#* Row 1, Column 2 is Apples
print(f'Cell {c.coordinate} is {c.value}')
#* Cell B1 is Apples
sheet['C1'].value
#* 73


sheet.cell(row=1, column=2)
#* <Cell 'Sheet1'.B1>
sheet.cell(row=1, column=2).value
#* 'Apples'

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
#* 1 Apples
#* 3 Pears
#* 5 Apples
#* 7 Strawberries

sheet.max_row
#* 7
sheet.max_column
#* 3

get_column_letter(1)
#* 'A'
get_column_letter(900)
#* 'AHP'
get_column_letter(sheet.max_column)
#* 'C'
column_index_from_string('AA')
#* 27

# Getting Rows and Columns from the Sheets
tuple(sheet['A1':'C3'])
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(f'{cellObj.coordinate} - {cellObj.value}')

#* A1 - 4/5/2015 1:34:02 PM
#* B1 - Apples
#* C1 - 73
#* A2 - 4/5/2015 3:41:23 AM
#* B2 - Cherries
#* C2 - 85
#* A3 - 4/6/2015 12:46:51 PM
#* B3 - Pears
#* C3 - 14

# create and save Excel document
wb = openpyxl.Workbook()
wb.get_sheet_names()
#* ['Sheet']

sheet = wb.active
sheet.title
#* 'Sheet'

sheet.title = 'New sheet name'
wb.get_sheet_names()
#* ['New sheet name']

wb.save('example_copy.xlsx')

# create and remove of sheets
wb.create_sheet(index=0, title='First Sheet')
#* <Worksheet "First Sheet">
wb.create_sheet(title='Last Sheet')
#* <Worksheet "Last Sheet">

wb.get_sheet_names()
#* ['First Sheet', 'New sheet name', 'Last Sheet']

wb.remove(wb.get_sheet_by_name('Last Sheet'))

wb.get_sheet_names()
#* ['First Sheet', 'New sheet name']

# writing to cell
sheet = wb.get_sheet_by_name('First Sheet')
sheet['A1'] = 'Hello world!'
sheet['A1'].value
#* 'Hello world!'

wb.save('example_copy.xlsx')

# setting the font style
wb = openpyxl.Workbook()
sheet = wb['Sheet']

fontObj1 = Font(name='Times New Roman', bold=True)
sheet['A1'].font = fontObj1
sheet['A1'] = 'Bold Times New Roman'

fontObj2 = Font(size=24, italic=True)
sheet['B1'].font = fontObj2
sheet['B1'] = '24 pt Italic'

wb.save('styles.xlsx')

# setting row height and column width
wb = openpyxl.Workbook()
sheet = wb.active

sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'

# the row height can be set to an integer or float value between 0 and 409.
# the default row height is 12.75
sheet.row_dimensions[1].height = 50

# the column width can be set to an integer or float value between 0 and 255.
# the default column width is 8.43 characters. 
sheet.column_dimensions['B'].width = 30

wb.save('dimensions.xlsx')

# merge cells
wb = openpyxl.Workbook()
sheet = wb.active

sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'

sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'

wb.save('merged.xlsx')

# unmerge cells
wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active

sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')

wb.save('merged.xlsx')

# freeze panes
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active

# freeze_panes setting      | Rows and columns frozen
# sheet.freeze_panes = 'A2' | Row 1
# sheet.freeze_panes = 'B1' | Column A
# sheet.freeze_panes = 'C1' | Columns A and B
# sheet.freeze_panes = 'C2' | Row 1 and columns A and B
sheet.freeze_panes = 'A2'

wb.save('freezeExample.xlsx')

# formulas
wb = openpyxl.Workbook()
sheet = wb.active

sheet['A1'] = 200
sheet['A2'] = 300

sheet['A3'] = '=SUM(A1:A2)'

sheet['A3'].value
#* '=SUM(A1:A2)'

wb.save('writeFormula.xlsx')

# # create a file with test data
# import excel_file
# excel_file.create_file('moving.xlsx')

# inserting/deleting and moving rows, columns and ranges
wb = openpyxl.load_workbook('moving.xlsx')
sheet = wb.active

# add 2 rows between rows 6 and 7 
sheet.insert_rows(7, 2) # delete_rows()

# delete the columns E:H
sheet.delete_cols(5, 4) # insert_cols()

# moving range of cells
sheet.move_range("A9:D15", rows=-2) # two rows up

# moving a range of cells with formula update
sheet.move_range("A1:D12", rows=1, cols=1, translate=True) 

wb.save('moving.xlsx')